#!/usr/bin/env python3
"""
Generate chillibyte-project-trackers.xlsx — all operational trackers in one
workbook, one tab each.

Run:  build/venv/bin/python build/make_xlsx.py

Nine tabs come from the existing Google Sheet. Two are new:
  - Estimate baseline   (G1 — nothing measured change against before)
  - Variation register  (G2 — the change log had no commercial arm)
The Change log tab also gains effort/chargeable/variation columns.
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

OUT = os.path.join(os.path.dirname(__file__), '..', 'templates', 'xlsx')

DARK = '212333'
RED = 'ED1B24'
GREEN = '8DC53E'
GREY = 'F5F5F5'
BEIGE = 'F4F2ED'

HDR_FONT = Font(name='Aptos', size=10, bold=True, color='FFFFFF')
HDR_FILL = PatternFill('solid', fgColor=DARK)
NEW_FILL = PatternFill('solid', fgColor='4A6B22')      # new columns, still readable white-on-dark
BODY_FONT = Font(name='Aptos', size=10)
HINT_FONT = Font(name='Aptos', size=9, italic=True, color='6A6A6A')
NOTE_FONT = Font(name='Aptos', size=9, italic=True, color='6A6A6A')
TITLE_FONT = Font(name='Aptos', size=12, bold=True, color=DARK)

THIN = Side(style='thin', color='D9D9D9')
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS = '"Not started,In progress,Blocked,Complete,N/A"'
APPROVAL = '"Draft,In review,Approved,Superseded"'
YESNO = '"Yes,No"'
DECISION = '"Proposed,Accepted,Rejected,Deferred"'
CHG_DECISION = '"Assess,Accepted,Rejected,Deferred"'
SEVERITY = '"Blocker,Major,Minor,Cosmetic"'
LIKELIHOOD = '"Low,Medium,High"'


def sheet(wb, name, title, note, columns, example=None, new_cols=(),
          validations=None, rows=40, widths=None):
    """One tracker tab. Row 1 = title, row 2 = note, row 3 = headers,
    row 4 = greyed example, rows 5+ = blank and validated."""
    ws = wb.create_sheet(name)

    ws['A1'] = title
    ws['A1'].font = TITLE_FONT
    ws['A2'] = note
    ws['A2'].font = NOTE_FONT

    for i, col in enumerate(columns, start=1):
        c = ws.cell(row=3, column=i, value=col)
        c.font = HDR_FONT
        c.fill = NEW_FILL if col in new_cols else HDR_FILL
        c.alignment = Alignment(vertical='center', wrap_text=True)
        c.border = BORDER

    if example:
        for i, v in enumerate(example, start=1):
            c = ws.cell(row=4, column=i, value=v)
            c.font = HINT_FONT
            c.fill = PatternFill('solid', fgColor=BEIGE)
            c.alignment = Alignment(vertical='top', wrap_text=True)
            c.border = BORDER

    first_blank = 5 if example else 4
    for r in range(first_blank, first_blank + rows):
        for i in range(1, len(columns) + 1):
            c = ws.cell(row=r, column=i)
            c.font = BODY_FONT
            c.border = BORDER
            c.alignment = Alignment(vertical='top', wrap_text=True)

    # Column widths
    for i, col in enumerate(columns, start=1):
        letter = get_column_letter(i)
        if widths and col in widths:
            ws.column_dimensions[letter].width = widths[col]
        else:
            ws.column_dimensions[letter].width = max(12, min(38, len(col) + 8))

    ws.row_dimensions[3].height = 32
    ws.freeze_panes = ws.cell(row=first_blank, column=1)
    ws.auto_filter.ref = f'A3:{get_column_letter(len(columns))}{first_blank + rows - 1}'

    # Dropdowns
    if validations:
        for col_name, formula in validations.items():
            if col_name not in columns:
                continue
            idx = columns.index(col_name) + 1
            letter = get_column_letter(idx)
            dv = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
            dv.error = 'Pick a value from the list'
            dv.errorTitle = 'Invalid entry'
            ws.add_data_validation(dv)
            dv.add(f'{letter}{first_blank}:{letter}{first_blank + rows - 1}')

    return ws


def fill_checklist(ws, rows, ncols, status_col, columns, validations,
                   dep_col=None, extra_blank=10):
    """Write pre-filled checklist rows from row 4, then blank rows for additions,
    and re-apply the filter range and dropdowns over the whole thing. The generic
    sheet() helper sizes itself for empty tabs, so these two need fixing up."""
    start = 4
    for i, row in enumerate(rows, start=start):
        ws.cell(row=i, column=1, value=row[0]).font = BODY_FONT
        ws.cell(row=i, column=2, value=row[1]).font = BODY_FONT
        ws.cell(row=i, column=status_col, value='Not started').font = BODY_FONT
        if dep_col and len(row) > 2:
            ws.cell(row=i, column=dep_col, value=row[2]).font = BODY_FONT

    last = start + len(rows) + extra_blank - 1
    for r in range(start, last + 1):
        for col in range(1, ncols + 1):
            c = ws.cell(row=r, column=col)
            c.border = BORDER
            c.alignment = Alignment(vertical='top', wrap_text=True)
            if c.font is None or c.font.name != 'Aptos':
                c.font = BODY_FONT

    ws.auto_filter.ref = f'A3:{get_column_letter(ncols)}{last}'

    for col_name, formula in validations.items():
        if col_name not in columns:
            continue
        letter = get_column_letter(columns.index(col_name) + 1)
        dv = DataValidation(type='list', formula1=formula, allow_blank=True, showDropDown=False)
        ws.add_data_validation(dv)
        dv.add(f'{letter}{start}:{letter}{last}')


def build():
    wb = Workbook()
    wb.remove(wb.active)

    # --- README ------------------------------------------------------------
    ws = wb.create_sheet('Read me')
    ws['A1'] = 'Chillibyte website project trackers'
    ws['A1'].font = Font(name='Aptos', size=16, bold=True, color=DARK)
    lines = [
        '',
        'One workbook per project. Duplicate it, rename it, delete the tabs the project does not need.',
        'The spec builder tells you which tabs this project needs — and which it does not.',
        '',
        'Row 3 on every tab is the header. Row 4 is a greyed worked example — delete it or type over it.',
        'Status columns have dropdowns. Panes are frozen and filters are on.',
        '',
        'TABS',
        '  Gate tracker          The spine. One row per gate. Start here.',
        '  Estimate baseline     NEW. The number every later change is measured against.',
        '  Decision log          Why something was chosen. Protects reasoning.',
        '  Change log            What moved after approval. Protects scope.',
        '  Variation register    NEW. What the change was worth. Protects revenue.',
        '  Sitemap               Pages, URLs, purpose, approval status.',
        '  Page inventory        The same pages as a work plan.',
        '  Content matrix        Every content item, with an owner AND a deadline.',
        '  Block inventory       Reuse / adapt / build new. Controls development scope.',
        '  Risks & assumptions   With owners and mitigations.',
        '  QA                    Tested against the agreed support matrix, not against everything.',
        '  Launch                Pre- and post-launch checklist.',
        '',
        'THE THREE RECORDS PEOPLE CONFUSE',
        '  Decision log  — why we chose something. Stops the same argument happening three times.',
        '  Change log    — what changed after approval, and its impact.',
        '  Variation     — what that change is worth, agreed with the client BEFORE the work happens.',
        '',
        'A change log without a variation register just documents unpaid work politely.',
        '',
        'You can still choose to absorb something. The process does not require charging for',
        'everything — it requires the decision to be visible.',
    ]
    for i, t in enumerate(lines, start=2):
        ws.cell(row=i, column=1, value=t).font = (
            Font(name='Aptos', size=10, bold=True, color=RED) if t.isupper() and t
            else Font(name='Aptos', size=10)
        )
    ws.column_dimensions['A'].width = 110

    # --- Gate tracker -------------------------------------------------------
    sheet(wb, 'Gate tracker', 'Gate tracker',
          'The spine of the project. Approval names artefacts AND their version — otherwise there is '
          'nothing to point at later.',
          ['Gate', 'Name', 'Route', 'Required artefacts (with version)', 'Owner', 'Approver',
           'Basecamp ticket', 'Status', 'Sent for approval', 'Approved date', 'Days taken',
           'Within SLA?', 'Decision log refs', 'Change log refs', 'Notes'],
          example=['0', 'Sales handover accepted', 'All', 'Sales handover note v1.0', 'Sales',
                   'Head of Web', 'Gate 0: sales handover accepted', 'Complete', '14/04', '15/04',
                   '1', 'Yes', '', '', 'Configurator excluded — phase 2 candidate'],
          new_cols=('Sent for approval', 'Days taken', 'Within SLA?'),
          validations={'Status': STATUS, 'Within SLA?': YESNO},
          widths={'Required artefacts (with version)': 40, 'Notes': 34, 'Basecamp ticket': 30},
          rows=24)

    # --- Estimate baseline (NEW) -------------------------------------------
    sheet(wb, 'Estimate baseline', 'Estimate baseline  [NEW]',
          'Set at Gate 1, re-baselined at Gate 10, actuals recorded at Gate 14. Without this, scope '
          'creep cannot be demonstrated — only asserted.',
          ['Phase', 'Step', 'Gate 1 baseline (hrs)', 'Gate 10 re-baseline (hrs)', 'Actual (hrs)',
           'Variance', 'Variance %', 'Of which billed as variation', 'Why it varied'],
          example=['Structure', 'Block identification', '4', '4', '5', '+1', '25%', '0',
                   'One extra block found during outlining'],
          new_cols=('Gate 1 baseline (hrs)', 'Gate 10 re-baseline (hrs)', 'Actual (hrs)',
                    'Variance', 'Variance %', 'Of which billed as variation', 'Why it varied'),
          widths={'Why it varied': 42, 'Step': 30},
          rows=30)

    # --- Decision log -------------------------------------------------------
    sheet(wb, 'Decision log', 'Decision log',
          'Why something was chosen. Protects reasoning — and lets someone new understand the project '
          'without asking anyone.',
          ['Decision ID', 'Date', 'Project phase', 'Decision title', 'Context', 'Options considered',
           'Selected decision', 'Rationale', 'Impact', 'Reversal cost', 'Owner', 'Approver',
           'Approval source', 'Status', 'Related artefacts', 'Notes'],
          example=['DEC-001', '', 'Discovery', 'Project route selected',
                   'Discovery indicates likely complexity', 'Light; standard; full', 'Standard',
                   'Default route for modular B2B service sites', 'Sets artefact depth and gate count',
                   'Medium', 'Head of Web', 'Internal', 'Basecamp Gate 1', 'Proposed',
                   'Discovery summary', ''],
          validations={'Status': DECISION, 'Reversal cost': LIKELIHOOD},
          widths={'Context': 30, 'Options considered': 28, 'Rationale': 34, 'Impact': 30, 'Notes': 26},
          rows=30)

    # --- Change log ---------------------------------------------------------
    sheet(wb, 'Change log', 'Change log',
          'What moved after approval. The last five columns are new: a change process that cannot raise '
          'a charge just converts scope creep into unpaid work more politely.',
          ['Change ID', 'Date raised', 'Raised by', 'Description', 'Related approved item',
           'Classification', 'Reason', 'Impact area', 'Impact detail', 'Effort (hrs)',
           'Chargeable?', 'Variation ref', 'Decision', 'Client approved', 'Basecamp reference',
           'Status', 'Completed date', 'Notes'],
          example=['CHG-001', '', 'Client', 'Add new page after sitemap approval', 'Sitemap v1.0',
                   'New scope', 'New stakeholder request', 'Structure / content / build',
                   'New URL, navigation, copy and possible template work', '6', 'Yes', 'VAR-001',
                   'Accepted', 'Yes', '', 'Open', '', ''],
          new_cols=('Classification', 'Effort (hrs)', 'Chargeable?', 'Variation ref', 'Client approved'),
          validations={'Decision': CHG_DECISION, 'Chargeable?': YESNO, 'Client approved': YESNO,
                       'Status': '"Open,In progress,Closed,Deferred"',
                       'Classification': '"Correction,Preference,New scope"'},
          widths={'Description': 34, 'Impact detail': 34, 'Notes': 24},
          rows=30)

    # --- Variation register (NEW) ------------------------------------------
    sheet(wb, 'Variation register', 'Variation register  [NEW]',
          'What each accepted change was worth. This tab is the difference between recording scope creep '
          'and being paid for it.',
          ['Variation ref', 'Date', 'Change log ref', 'Description', 'Effort (hrs)', 'Rate',
           'Value', 'Offered as', 'Client decision', 'Decision date', 'Schedule impact (days)',
           'Invoiced?', 'Notes'],
          example=['VAR-001', '', 'CHG-001', 'Additional service line page and template variant',
                   '6', '£86.67', '£520', 'Proceed now / defer / decline', 'Accepted', '', '2',
                   'No', 'Baseline updated'],
          new_cols=('Variation ref', 'Date', 'Change log ref', 'Description', 'Effort (hrs)', 'Rate',
                    'Value', 'Offered as', 'Client decision', 'Decision date',
                    'Schedule impact (days)', 'Invoiced?', 'Notes'),
          validations={'Client decision': '"Proceed,Defer,Decline,Absorbed"', 'Invoiced?': YESNO},
          widths={'Description': 40, 'Offered as': 28, 'Notes': 26},
          rows=25)

    # --- Sitemap ------------------------------------------------------------
    sheet(wb, 'Sitemap', 'Sitemap',
          'Approved at Gate 3. Where there is an existing site, every old URL maps to a new one or to a '
          'documented retirement.',
          ['Page', 'Proposed URL', 'Parent', 'Purpose', 'Primary audience', 'Primary CTA',
           'Secondary CTA', 'Template type', 'Navigation placement', 'SEO priority',
           'Old URL', 'Redirect type', 'Content owner', 'Approval status', 'Notes'],
          example=['Home', '/', 'n/a', 'Introduce business and route users', 'All audiences',
                   'Enquire', 'View services', 'Unique', 'Main nav', 'High', '/', '—',
                   'Copywriter / client', 'Draft', ''],
          new_cols=('Old URL', 'Redirect type'),
          validations={'Approval status': APPROVAL,
                       'Template type': '"Unique,Repeatable,Landing,Standard,Index"',
                       'Redirect type': '"301,410,None — new page,Unchanged"',
                       'SEO priority': LIKELIHOOD},
          widths={'Purpose': 34, 'Notes': 24},
          rows=45)

    # --- Page inventory -----------------------------------------------------
    sheet(wb, 'Page inventory', 'Page inventory',
          'The same pages as a work plan. Template count drives effort far more than page count does.',
          ['Page', 'URL', 'Page type', 'Unique or repeatable', 'Primary job', 'Audience',
           'Content status', 'Design status', 'Build status', 'Template / block pattern',
           'Media required', 'Owner', 'Approval status', 'Notes'],
          example=['Home', '/', 'Landing', 'Unique', 'Set proposition and route users', 'All',
                   'Not started', 'Not started', 'Not started', 'Home pattern',
                   'Hero image / service imagery', 'Head of Web', 'Draft', ''],
          validations={'Content status': STATUS, 'Design status': STATUS, 'Build status': STATUS,
                       'Approval status': APPROVAL,
                       'Unique or repeatable': '"Unique,Repeatable"'},
          widths={'Primary job': 32, 'Notes': 24},
          rows=45)

    # --- Content matrix -----------------------------------------------------
    sheet(wb, 'Content matrix', 'Content matrix',
          'Every content item has an owner AND a deadline. Late client content is the most common cause '
          'of website overrun — the deadline column is the fix.',
          ['Page', 'Section', 'Content needed', 'Approx. length', 'Content owner', 'Source',
           'Deadline', 'Days late', 'Escalated?', 'Media needed', 'Media owner', 'Status',
           'Dependency', 'Notes'],
          example=['Home', 'Hero', 'Headline, intro, CTA text', '30–60 words', 'Copywriter',
                   'Discovery summary', '', '', 'No', 'Hero image', 'Client', 'Not started',
                   'Style direction', ''],
          new_cols=('Deadline', 'Days late', 'Escalated?'),
          validations={'Status': STATUS, 'Escalated?': YESNO},
          widths={'Content needed': 32, 'Notes': 24},
          rows=50)

    # --- Block inventory ----------------------------------------------------
    sheet(wb, 'Block inventory', 'Block inventory',
          'Approved at Gate 5. Collapse duplicates hard — if this list is as long as your section list, '
          'no consolidation has happened.',
          ['Block', 'Used on', 'Status', 'Reuse decision', 'Purpose', 'Fields', 'Design needed',
           'Responsive notes', 'PHP structure', 'field group', 'Base SCSS', 'Theme SCSS',
           'Design library entry', 'Build impact', 'Assigned to', 'Est. hrs', 'Actual hrs',
           'Handed back', 'Approval status', 'Notes'],
          example=['Hero', 'Home, landing pages', 'Existing / adapt', 'Adapt existing',
                   'Introduce page and direct action', 'Heading, intro, image, buttons', 'Yes',
                   'Mobile image crop needs checking', 'Reusable block PHP', 'Hero group',
                   'Core layout', 'Project styling', 'Yes', 'Medium', 'Web Developer', '5', '',
                   'No', 'Draft', ''],
          new_cols=('Assigned to', 'Est. hrs', 'Actual hrs', 'Handed back'),
          validations={'Reuse decision': '"Reuse,Adapt existing,Build new,One-off"',
                       'Build impact': LIKELIHOOD, 'Approval status': APPROVAL,
                       'Design needed': '"Yes,Minor,No"',
                       'Design library entry': YESNO},
          widths={'Purpose': 30, 'Fields': 30, 'Responsive notes': 30, 'Notes': 22},
          rows=35)

    # --- Risks & assumptions ------------------------------------------------
    sheet(wb, 'Risks & assumptions', 'Risks and assumptions',
          'Every entry needs an owner and a validation route. A risk with no owner is a wish.',
          ['ID', 'Type', 'Date', 'Description', 'Impact if realised', 'Likelihood', 'Impact',
           'Owner', 'Mitigation / validation', 'Status', 'Related decision', 'Notes'],
          example=['RA-001', 'Assumption', '', 'Client will supply usable photography',
                   'Design quality and launch timing affected', 'Medium', 'Medium', 'Account Management',
                   'Confirm asset availability during discovery', 'Open', '', ''],
          validations={'Type': '"Risk,Assumption"', 'Likelihood': LIKELIHOOD, 'Impact': LIKELIHOOD,
                       'Status': '"Open,Mitigated,Realised,Closed"'},
          widths={'Description': 36, 'Impact if realised': 34, 'Mitigation / validation': 36},
          rows=30)

    # --- QA -----------------------------------------------------------------
    qa_rows = [
        ('Structure', 'Sitemap matches approved structure'),
        ('Navigation', 'Main navigation works across the support matrix'),
        ('Navigation', 'Subnav and breadcrumbs correct'),
        ('Content', 'No placeholder or dummy content remains'),
        ('Content', 'Headings match the approved page outlines'),
        ('Forms', 'Every form submits AND arrives at the correct destination'),
        ('Forms', 'Validation and error states work'),
        ('Responsive', 'Key pages checked at every breakpoint in the support matrix'),
        ('Responsive', 'Checked on every browser in the support matrix — and nothing beyond it'),
        ('Accessibility', 'Heading order is sequential, no skipped levels'),
        ('Accessibility', 'All text meets the contrast target'),
        ('Accessibility', 'Focus states visible on every interactive element'),
        ('Accessibility', 'Keyboard navigation reaches everything, including scroll regions'),
        ('Accessibility', 'Images have meaningful alt text'),
        ('Performance', 'Images optimised and correctly sized'),
        ('Performance', 'Core Web Vitals checked on the heaviest pages'),
        ('SEO', 'Titles and meta descriptions present and unique'),
        ('SEO', 'Indexability correct — staging noindex removed'),
        ('SEO', 'XML sitemap generated and correct'),
        ('Redirects', 'Every old URL tested against the current-state audit list'),
        ('Redirects', 'PDF and asset paths preserved or redirected'),
        ('Compliance', 'Cookie consent actually gates non-essential cookies'),
        ('Compliance', 'Privacy policy describes the site as built'),
        ('Compliance', 'Form data destination and retention documented'),
        ('Analytics', 'Analytics firing and consent mode configured'),
    ]
    ws = sheet(wb, 'QA', 'QA checklist',
               'Tested against the support matrix and accessibility target agreed at Gate 1 — and '
               'nothing beyond them. That is what makes QA finishable.',
               ['Category', 'Check', 'Owner', 'Status', 'Issue', 'Severity', 'Fix required',
                'Retested', 'Basecamp reference', 'Notes'],
               new_cols=('Retested',),
               validations=None,   # applied by fill_checklist over the real range
               widths={'Check': 52, 'Issue': 30, 'Notes': 24},
               rows=0)
    fill_checklist(ws, qa_rows, ncols=10, status_col=4, extra_blank=10,
                   validations={'Status': STATUS, 'Severity': SEVERITY,
                                'Fix required': YESNO, 'Retested': YESNO},
                   columns=['Category', 'Check', 'Owner', 'Status', 'Issue', 'Severity',
                            'Fix required', 'Retested', 'Basecamp reference', 'Notes'])

    # --- Launch -------------------------------------------------------------
    launch_rows = [
        ('Pre-launch', 'Final client approval captured (Gate 12)', 'Gate 12'),
        ('Pre-launch', 'Content freeze declared and confirmed to the client', 'Gate 12'),
        ('Pre-launch', 'Full backup taken AND restore verified', 'Hosting access'),
        ('Pre-launch', 'Redirects implemented and tested against the audit URL list', 'Current-state audit'),
        ('Pre-launch', 'DNS access confirmed and TTL reduced', 'DNS credentials'),
        ('Pre-launch', 'Hosting environment provisioned and tested', ''),
        ('Pre-launch', 'Integrations moved from sandbox to production', 'Credentials'),
        ('Pre-launch', 'Compliance cleared — consent, privacy, retention', 'Gate C5'),
        ('Pre-launch', 'Rollback plan written and read by whoever is launching', ''),
        ('Pre-launch', 'Launch window scheduled in working hours, not a Friday', ''),
        ('Pre-launch', 'Warranty window agreed with the client IN WRITING', 'Gate 13'),
        ('Launch', 'Deploy to production', 'Gate 13'),
        ('Launch', 'DNS switched', ''),
        ('Post-launch', 'Live smoke test — all key journeys walked', 'DNS propagated'),
        ('Post-launch', 'Every form submitted with real data and confirmed arriving', ''),
        ('Post-launch', 'Analytics and Search Console verified firing', ''),
        ('Post-launch', 'Sitemap submitted; crawl errors monitored for one week', ''),
        ('Post-launch', 'Redirects spot-checked live', ''),
        ('Post-launch', 'Client notified and handover notes delivered', ''),
        ('Post-launch', 'Warranty start date recorded', ''),
        ('Post-launch', 'Estimate variance recorded on the Estimate baseline tab', ''),
        ('Post-launch', 'Reusable blocks committed to the library', ''),
        ('Post-launch', 'Retrospective held — each action with an owner and a date', ''),
        ('Post-launch', 'Care plan conversation held while goodwill is high', ''),
    ]
    ws = sheet(wb, 'Launch', 'Launch checklist',
               'Most launch problems are something not prepared or something not tested. '
               'Preparation is the entire job.',
               ['Category', 'Task', 'Owner', 'Status', 'Dependency', 'Issue', 'Severity',
                'Basecamp reference', 'Notes'],
               validations=None,   # applied by fill_checklist over the real range
               widths={'Task': 56, 'Notes': 24},
               rows=0)
    fill_checklist(ws, [(c, t, d) for c, t, d in launch_rows], ncols=9, status_col=4,
                   dep_col=5, extra_blank=10,
                   validations={'Status': STATUS, 'Severity': SEVERITY},
                   columns=['Category', 'Task', 'Owner', 'Status', 'Dependency', 'Issue',
                            'Severity', 'Basecamp reference', 'Notes'])

    os.makedirs(OUT, exist_ok=True)
    path = os.path.abspath(os.path.join(OUT, 'chillibyte-project-trackers.xlsx'))
    wb.save(path)
    print(f'  chillibyte-project-trackers.xlsx  ({len(wb.sheetnames)} tabs)')
    for n in wb.sheetnames:
        print(f'    - {n}')


if __name__ == '__main__':
    print('Generating tracker workbook:')
    build()
    print('Done.')
